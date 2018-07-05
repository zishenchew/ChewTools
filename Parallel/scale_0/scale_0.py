# encoding: utf-8
########################################################################################################################
#
#    Parallel scaling tool
#    Author: Chew
#
#
#    Tool for adjusting the scale values of
#    A scene management tool for managing SER files on SVN
#
########################################################################################################################

import pymel.core as pm
import openpyxl as pyxl
import alphanum

class ParallelScale():

    def __init__(self):

        #pulling data from the excel file
        wb = pyxl.load_workbook(filename=r'\\p.sv\Tool\PrismTools\Parallel\ref\excelData.xlsx')
        ws = wb[wb.sheetnames[0]]

        self.nameScale = {}
        for cell in ws['C']:
            if cell.value and cell.value != u'キャラクタ':
                # print i.row
                self.nameScale[str(ws['A' + str(cell.row)].value) + ' ' + cell.value] = ws['H' + str(cell.row)].value
        self.nameList = []
        for i in self.nameScale:
            self.nameList.append(i)

        self.nameList.sort(alphanum.alphanum)


    def ParaUI(self):
        # setting basic UI height and width
        winWidth = 500
        winHeight = 100
        # checking for duplicate windows
        windowID = 'ParaScale'
        if pm.window(windowID, exists=True):
            pm.deleteUI(windowID)
        # re-adjusting UI if UI is not correct
        try:
            if pm.windowPref(windowID, q=True, height=True) != winHeight or pm.windowPref(windowID, q=True,
                                                                                          width=True) != winWidth:
                pm.windowPref(windowID, remove=True)
        except:
            pass

        pm.window(windowID, title=u'Parallel scale | Ver 2018 / 06 / 22', widthHeight=(winWidth, winHeight))
        masterFrame = pm.frameLayout(label = u'PARALLEL Scaling', labelIndent = 5, marginHeight = 5, nch = 5)
        row1 = pm.rowLayout('row1', nc=5, width=600, parent = masterFrame)
        pm.text(label=u'キャラ名：', width = 80, align = 'right')
        self.charalist = pm.optionMenu('listofchara', parent = row1)
        for i in self.nameList:
            pm.menuItem(label=i, parent = self.charalist)
        pm.text('', width = 10, parent = row1)
        pm.button(label = u'スケール', parent = row1, command = self.scaleChara)

        pm.showWindow()


    def scaleChara(self, mayaFalse):

        #this part does the importing and admin stuff
        try:
            pm.createReference('//p.sv/Prism/project/Parallel/element/character_Roll/scenes/' + pm.optionMenu(self.charalist, q = True, value = True)[-4:-1] + '001_SP01.mb',
                           namespace=':')
        except:
            pm.confirmDialog(title=u'Parallel scaling', message=u'Parallel フォルダーの許可がないです。')
        #this part does the scaling
        pm.optionMenu('listofchara', q = True, value = True)
        inverseScale = 1/self.nameScale[pm.optionMenu(self.charalist, q = True, value = True)]
        pm.scale('Reference|Root', inverseScale, inverseScale, inverseScale)
    
        for i in pm.listRelatives('Reference', ad = True, type = 'joint'):
            try:
                pm.copyKey(i)
                pm.pasteKey(i.replace('Reference', 'Reference1'))
            except:
                print 'no keys to copypasta'
        pm.scaleKey('Reference1|Root|Hips', valueScale = inverseScale, valuePivot = 0, attribute = 'translate')

        #this part cleans up the file
        pm.delete('Reference')
        pm.listReferences()[0].importContents(removeNamespace=False)
        pm.rename('Reference1', 'Reference')
